#!/usr/bin/env python3
"""
Organiza LISTADO DE PRODUCTOS.xlsx:
- Separa CLIENTE en nombre_cliente + rif
- Hoja CLIENTES (únicos)
- Hoja PRODUCTOS (lista limpia para Axones)
- Hoja INSTRUCCIONES
"""

from __future__ import annotations

import re
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "LISTADO DE PRODUCTOS.xlsx"
DEFAULT_OUTPUT = ROOT / "LISTADO DE PRODUCTOS - ORGANIZADO.xlsx"

HEADER_ROW = 6
DATA_START = 7

HEADER_FILL = PatternFill("solid", fgColor="FFC000")
HEADER_FONT = Font(bold=True, color="000000")
TITLE_FONT = Font(bold=True, size=12)


def cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def normalize_rif(raw: str) -> str:
    s = raw.upper().strip()
    s = re.sub(r"\s+", "", s)
    s = s.replace("RIF", "")
    s = s.replace(".", "").replace("_", "")
    s = s.replace("-", "")

    m = re.match(r"^([JVEGPC])(\d{7,9})$", s)
    if not m:
        return raw.strip()

    letter, digits = m.group(1), m.group(2)
    main, dv = digits[:-1], digits[-1]
    if letter == "J":
        return f"J-{main}-{dv}"
    return f"{letter}{main}{dv}"


def parse_cliente(raw: str) -> tuple[str, str]:
    text = raw.strip()
    if not text:
        return "", ""

    m = re.search(r"\(([^)]+)\)\s*$", text)
    if not m:
        return text, ""

    inner = m.group(1).strip()
    name = text[: m.start()].strip().rstrip(",").strip()

    rif_part = re.sub(r"^RIF\s*", "", inner, flags=re.I).strip()
    rif = normalize_rif(rif_part) if rif_part else ""
    return name, rif


def read_source_rows(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, str]] = []

    for r in range(DATA_START, ws.max_row + 1):
        producto = cell_str(ws.cell(r, 1).value)
        if not producto:
            continue

        cliente_raw = cell_str(ws.cell(r, 2).value)
        nombre_cliente, rif = parse_cliente(cliente_raw)
        cpe = cell_str(ws.cell(r, 3).value)
        mps = cell_str(ws.cell(r, 4).value)
        cod_barra = cell_str(ws.cell(r, 5).value)

        rows.append(
            {
                "fila_origen": str(r),
                "producto": producto,
                "cliente_original": cliente_raw,
                "nombre_cliente": nombre_cliente,
                "rif": rif,
                "cpe": "" if cpe.upper() == "N/A" else cpe,
                "mps": "" if mps.upper() == "N/A" else mps,
                "cod_barra": "" if cod_barra.upper() == "N/A" else cod_barra,
            }
        )

    return rows


def style_header_row(ws, row: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_text_columns(ws, col_indexes: list[int], start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in col_indexes:
            cell = ws.cell(r, c)
            cell.number_format = "@"
            if cell.value is not None and cell.value != "":
                cell.value = str(cell.value)


def autosize_columns(ws, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = max(len(cell_str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def build_clients(product_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by_rif: OrderedDict[str, dict[str, str]] = OrderedDict()
    for row in product_rows:
        rif = row["rif"]
        name = row["nombre_cliente"]
        if not rif and not name:
            continue
        key = rif or name.upper()
        if key not in by_rif:
            by_rif[key] = {"nombre_cliente": name, "rif": rif, "productos": "0"}
        by_rif[key]["productos"] = str(int(by_rif[key]["productos"]) + 1)
    return list(by_rif.values())


def write_instructions_sheet(ws) -> None:
    ws.title = "INSTRUCCIONES"
    lines = [
        "LISTADO DE PRODUCTOS — Guía para cargar en Axones",
        "",
        "Este archivo fue generado automáticamente a partir del listado original.",
        "",
        "HOJAS:",
        "  • CLIENTES — una fila por empresa (sin repetir). Cargar PRIMERO en Axones.",
        "  • PRODUCTOS — una fila por especificación. Cargar DESPUÉS de los clientes.",
        "  • ORIGINAL — copia de referencia del archivo fuente.",
        "",
        "COLUMNAS EN PRODUCTOS:",
        "  producto      → Nombre del producto en Axones (Especificaciones)",
        "  rif_cliente   → Enlaza con la hoja CLIENTES (mismo RIF)",
        "  nombre_cliente→ Solo referencia; en Axones se elige el cliente del maestro",
        "  cpe           → C.P.E. (texto; no borrar ceros a la izquierda)",
        "  mps           → M.P.P.S.",
        "  cod_barra     → Código de barra maestro",
        "",
        "ORDEN DE CARGA EN AXONES:",
        "  1. Datos maestros → Clientes (nombre + RIF de hoja CLIENTES)",
        "  2. Datos maestros → Productos / Nueva especificación (por cada fila de PRODUCTOS)",
        "  3. Opcional: completar Tipo de impresión y Estructura en cada producto",
        "",
        "REGLAS:",
        "  • No puede haber dos productos con el mismo nombre para el mismo cliente.",
        "  • CPE y código de barra conviene tenerlos completos antes de crear órdenes de trabajo.",
        "  • Si falta CPE o barra (ej. fila TE ARRECIFE), completar manualmente después.",
        "",
        "MANTENIMIENTO DEL EXCEL:",
        "  • Agregar productos nuevos en hoja PRODUCTOS (no mezclar nombre + RIF en una celda).",
        "  • Si entra un cliente nuevo, agregarlo primero en CLIENTES y luego el producto.",
        "  • Formato Texto en cpe y cod_barra para evitar notación científica.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(i, 1, line)
        if i == 1:
            cell.font = TITLE_FONT
    ws.column_dimensions["A"].width = 92


def write_workbook(output: Path, product_rows: list[dict[str, str]], source_path: Path) -> None:
    clients = build_clients(product_rows)
    wb = openpyxl.Workbook()

    # CLIENTES
    ws_c = wb.active
    ws_c.title = "CLIENTES"
    c_headers = ["nombre_cliente", "rif", "cantidad_productos"]
    ws_c.append(c_headers)
    for c in clients:
        ws_c.append([c["nombre_cliente"], c["rif"], c["productos"]])
    style_header_row(ws_c, 1, len(c_headers))
    set_text_columns(ws_c, [2], 2, ws_c.max_row)

    # PRODUCTOS
    ws_p = wb.create_sheet("PRODUCTOS")
    p_headers = [
        "producto",
        "rif_cliente",
        "nombre_cliente",
        "cpe",
        "mps",
        "cod_barra",
        "fila_origen",
    ]
    ws_p.append(p_headers)
    for row in product_rows:
        ws_p.append(
            [
                row["producto"],
                row["rif"],
                row["nombre_cliente"],
                row["cpe"],
                row["mps"],
                row["cod_barra"],
                row["fila_origen"],
            ]
        )
    style_header_row(ws_p, 1, len(p_headers))
    set_text_columns(ws_p, [2, 4, 6], 2, ws_p.max_row)

    # ORIGINAL (reference)
    ws_o = wb.create_sheet("ORIGINAL")
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    src_ws = src_wb[src_wb.sheetnames[0]]
    for r in range(1, src_ws.max_row + 1):
        ws_o.append([cell_str(src_ws.cell(r, c).value) for c in range(1, 6)])

    # INSTRUCCIONES (primera pestaña)
    ws_i = wb.create_sheet("INSTRUCCIONES", 0)
    write_instructions_sheet(ws_i)

    for sheet in (ws_c, ws_p):
        autosize_columns(sheet)

    wb.save(output)


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not input_path.is_file():
        print(f"No se encontró: {input_path}", file=sys.stderr)
        return 1

    rows = read_source_rows(input_path)
    missing_rif = [r for r in rows if not r["rif"]]
    if missing_rif:
        print("Advertencia: filas sin RIF parseado:", file=sys.stderr)
        for r in missing_rif:
            print(f"  fila {r['fila_origen']}: {r['cliente_original']}", file=sys.stderr)

    write_workbook(output_path, rows, input_path)
    clients = build_clients(rows)
    print(f"OK: {len(rows)} productos, {len(clients)} clientes únicos")
    print(f"Guardado: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
